import tkinter as tk
from tkinter import filedialog
import cv2
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font
from PIL import Image



window = tk.Tk()
entry1 = tk.Entry(window, width=40)
entry2 = tk.Entry(window, width=40)


#FOR LOGIC 
def compare_images(image1, image2, output_path):
    # Read the images
    img1 = cv2.imread(image1)
    img2 = cv2.imread(image2)

    # Get dimession of img1
    height, width, _ = img1.shape

    resize_img2 = cv2.resize(img2, (width, height))

    # Convert the images to grayscale
    gray1 = cv2.cvtColor(img1, cv2.COLOR_BGR2GRAY)
    gray2 = cv2.cvtColor(resize_img2, cv2.COLOR_BGR2GRAY)

    # Compute the absolute difference between the two images
    diff = cv2.absdiff(gray1, gray2)

    # Threshold the difference image
    _, threshold = cv2.threshold(diff, 30, 255, cv2.THRESH_BINARY)

    # Find contours in the threshold image
    contours, _ = cv2.findContours(threshold, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    

    # Draw rectangles around the differences
    for contour in contours:
        x, y, w, h = cv2.boundingRect(contour)
        cv2.rectangle(resize_img2, (x, y), (x + w, y + h), (0, 0, 255), 2)

    # Save the output image
    cv2.imwrite(output_path, resize_img2)
    image = Image.open(output_path)
    image.show()

# FOR UI 
    
dimessions = {}
    
def get_image_dimensions(file_path):
    image = Image.open(file_path)
    width, height = image.size
    return width, height

def browse_file(entry, name):
    filename = filedialog.askopenfilename(filetypes=[("Image files", "*.png;*.jpg;*.jpeg")])
    if filename:
        width, height = get_image_dimensions(filename)
        dimessions[name] = {
            "width": width,
            "height": height 
        }
    entry.delete(0, tk.END)
    entry.insert(0, filename)
    writeUi()

def center_window(window, width, height):
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()

    x_coordinate = (screen_width - width) // 2
    y_coordinate = (screen_height - height) // 2

    window.geometry(f"{width}x{height}+{x_coordinate}+{y_coordinate}")

def writeUi(): 
    # Create the main window
    window.title("Happy Compare")
    center_window(window, 600, 170)

    # Create the first entry widget for file path 1
    label1 = tk.Label(window, text="Path old image:")
    label1.grid(row=0, column=0, padx=10, pady=10)

    entry1.grid(row=0, column=1, padx=10, pady=10)

    xyz1 = tk.Entry(window, state="disabled", justify="right")
    xyz1.grid(row=0, column=2, padx=5, pady=5)
    if 'image1' in dimessions:
        xyz1.config(state='normal')
        xyz1.delete(0, tk.END)
        xyz1.insert(0, f"{dimessions['image1']['width']}x{dimessions['image1']['height']}")
        xyz1.config(state='disabled')

    browse_button1 = tk.Button(window, text="Browse", command=lambda: browse_file(entry1, 'image1'))
    browse_button1.grid(row=0, column=3, padx=10, pady=10)

    # Create the second entry widget for file path 2
    label2 = tk.Label(window, text="Path new image:")
    label2.grid(row=1, column=0, padx=10, pady=10)
    
    entry2.grid(row=1, column=1, padx=10, pady=10)

    xyz2 = tk.Entry(window, state="readonly", justify="right")
    xyz2.grid(row=1, column=2, padx=5, pady=5)
    if 'image2' in dimessions:
        xyz2.config(state='normal')
        xyz2.delete(0, tk.END)
        xyz2.insert(0, f"{dimessions['image2']['width']}x{dimessions['image2']['height']}")
        xyz2.config(state='disabled')

    browse_button2 = tk.Button(window, text="Browse", command=lambda: browse_file(entry2, 'image2'))
    browse_button2.grid(row=1, column=3, padx=10, pady=10)

    # Create the button
    button = tk.Button(window, text="Compare", command=on_button_click)
    button.grid(row=2, column=0, columnspan=3, pady=10)

    # Start the Tkinter event loop
    window.mainloop()

def resize_image(file_path):
    original_image = Image.open(file_path)
    original_size = original_image.size
    new_size = (original_size[0] // 2, original_size[1] // 2)
    resized_image = original_image.resize(new_size)
    return resized_image



def writeIntoExcel(file_path1): 
    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active


    ws['A1'].value = 'Layout old'
    ws['A1'].font = Font(size=12, bold=True)

    start_cell = 'A2'
    end_cell = 'A40'

    #Image old

    img_old = ExcelImage(resize_image(file_path1))
    ws.add_image(img_old, 'A2')

    # Save the workbook
    wb.save('output.xlsx')

def on_button_click():
    # Retrieve the file paths from the entry widgets
    file_path1 = entry1.get()
    file_path2 = entry2.get()

    # Print the file paths (you can modify this part as needed)
    compare_images(file_path1, file_path2, './diff.png')

    writeIntoExcel(file_path1)
    


if __name__ == "__main__":
    writeUi()
